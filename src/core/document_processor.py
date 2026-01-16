"""
Document processing utilities for various file formats.
Supports PDF, DOCX, TXT, MD, and XLSX files.
"""

from typing import Dict, Any, List
from pathlib import Path
from loguru import logger
import pypdf
from docx import Document
import openpyxl
import markdown
from bs4 import BeautifulSoup


class DocumentProcessor:
    """Process various document formats and extract text with metadata."""

    @staticmethod
    def process_file(file_path: str) -> Dict[str, Any]:
        """
        Process a file and extract text content.

        Args:
            file_path: Path to the file

        Returns:
            Dictionary with text, metadata, and processing info
        """
        path = Path(file_path)
        extension = path.suffix.lower()

        try:
            if extension == ".pdf":
                return DocumentProcessor._process_pdf(file_path)
            elif extension == ".docx":
                return DocumentProcessor._process_docx(file_path)
            elif extension in [".txt", ".md"]:
                return DocumentProcessor._process_text(file_path)
            elif extension == ".xlsx":
                return DocumentProcessor._process_xlsx(file_path)
            else:
                raise ValueError(f"Unsupported file type: {extension}")

        except Exception as e:
            logger.error("Error processing file {}: {}", file_path, e)
            raise

    @staticmethod
    def _process_pdf(file_path: str) -> Dict[str, Any]:
        """Extract text from PDF with page numbers."""
        text_parts = []
        metadata = {
            "filename": Path(file_path).name,
            "file_type": "pdf",
            "pages": [],
        }

        with open(file_path, "rb") as f:
            pdf_reader = pypdf.PdfReader(f)
            total_pages = len(pdf_reader.pages)

            for page_num, page in enumerate(pdf_reader.pages, start=1):
                page_text = page.extract_text()
                if page_text.strip():
                    text_parts.append(f"[Seite {page_num}]\n{page_text}")
                    metadata["pages"].append({
                        "page": page_num,
                        "text_length": len(page_text),
                    })

            metadata["total_pages"] = total_pages

        return {
            "text": "\n\n".join(text_parts),
            "metadata": metadata,
        }

    @staticmethod
    def _process_docx(file_path: str) -> Dict[str, Any]:
        """Extract text from DOCX."""
        doc = Document(file_path)
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]

        return {
            "text": "\n\n".join(paragraphs),
            "metadata": {
                "filename": Path(file_path).name,
                "file_type": "docx",
                "paragraphs": len(paragraphs),
            },
        }

    @staticmethod
    def _process_text(file_path: str) -> Dict[str, Any]:
        """Extract text from TXT or MD files."""
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()

        # If markdown, convert to plain text
        if Path(file_path).suffix.lower() == ".md":
            html = markdown.markdown(content)
            soup = BeautifulSoup(html, "html.parser")
            content = soup.get_text()

        return {
            "text": content,
            "metadata": {
                "filename": Path(file_path).name,
                "file_type": Path(file_path).suffix.lower().lstrip("."),
            },
        }

    @staticmethod
    def _process_xlsx(file_path: str) -> Dict[str, Any]:
        """Extract text from XLSX (all sheets)."""
        workbook = openpyxl.load_workbook(file_path)
        text_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text = []

            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell else "" for cell in row)
                if row_text.strip():
                    sheet_text.append(row_text)

            if sheet_text:
                text_parts.append(f"[Tabelle: {sheet_name}]\n" + "\n".join(sheet_text))

        return {
            "text": "\n\n".join(text_parts),
            "metadata": {
                "filename": Path(file_path).name,
                "file_type": "xlsx",
                "sheets": workbook.sheetnames,
            },
        }
